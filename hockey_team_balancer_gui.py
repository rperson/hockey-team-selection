import pandas as pd
import random
from dataclasses import dataclass
from copy import deepcopy
import tkinter as tk
from typing import cast
from tkinter import filedialog, messagebox
import email
from email.message import EmailMessage # Added
from email import policy
import re
import string
from openpyxl.styles import Font, Alignment, Border, Side


# -----------------------------
# Player Class
# -----------------------------

@dataclass
class Player:
    name: str
    rank: int # The player's skill rank
    position: str # "F", "D", or "FLEX" (or "?" for unrecognized players)
    is_unrecognized: bool = False # New field to mark players not found in the Excel data


# -----------------------------
# Config
# -----------------------------

FORWARDS_TARGET = 6
DEFENCE_TARGET = 4
ITERATIONS = 7000

TEAM_A_NAME = "Light Team"
TEAM_B_NAME = "Dark Team"

TEAM_A_JERSEY = "LIGHT"
TEAM_B_JERSEY = "DARK"

OUTPUT_FILE = "game_night_teams.xlsx"


# -----------------------------
# Helpers
# -----------------------------

def is_selected(value):
    if pd.isna(value):
        return False
    return str(value).strip().upper() in ["TRUE", "1", "YES", "Y"]


def empty_team():
    return {"F": [], "D": [], "total": 0}


def can_add(team, pos):
    if pos == "F":
        return len(team["F"]) < FORWARDS_TARGET
    if pos == "D":
        return len(team["D"]) < DEFENCE_TARGET
    return False

def extract_players_from_eml(eml_file_path, player_data_lookup):
    """
    Parses an EML file to extract the list of player names from the final roster section.
    Names are validated against the provided player_data_lookup from the Excel file.
    """
    unrecognized_player_names = [] # List to store names not found in player_data_lookup
    with open(eml_file_path, 'rb') as fp: # type: ignore
        msg: EmailMessage = email.message_from_binary_file(fp, policy=policy.default)

    player_names = []
    payload = ""

    # Attempt to get the plain text part of the email
    if msg.is_multipart():
        for part in msg.iter_parts():
            if part.get_content_type() == 'text/plain':
                payload = part.get_payload(decode=True)
                break
        else:
            raise ValueError("No plain text part found in the EML file.")
    else:
        if msg.get_content_type() == 'text/plain':
            payload = msg.get_payload(decode=True)
        else:
            raise ValueError("EML file is not plain text or multipart with a plain text part.")

    # Compile the roster start regex once, outside the loop for efficiency
    # Using re.VERBOSE for better readability
    roster_start_pattern = re.compile(r"""
        .*          # Match any characters before "which"
        \bwhich\b   # Match the whole word "which"
        .*          # Match any characters in between
        \broster\b  # Match the whole word "roster"
        .*          # Match any characters in between
        \bof\b      # Match the whole word "of"
        .*          # Match any characters in between
        :           # Match the colon character
    """, re.IGNORECASE | re.VERBOSE)

    lines = payload.splitlines()

    in_roster_section = False
    for line in lines:
        stripped_line = line.decode().strip()

        if roster_start_pattern.search(stripped_line):
            in_roster_section = True
            continue # Skip the marker line itself

        if in_roster_section:
            if not stripped_line:
                # Allow blank lines within the roster section
                continue

            # Clean up potential extra spaces or non-breaking spaces
            clean_name = re.sub(r'[\s\xa0]+', ' ', stripped_line).strip()

            # Validate the name against the Excel player data
            normalized_clean_name = clean_name.upper()
            if normalized_clean_name in player_data_lookup:
                player_names.append(clean_name)
            else:
                # This non-empty line is not a recognized player name from the lookup,
                # but it might be the end of the roster.
                word_count = len(clean_name.split())
                has_punctuation = any(char in string.punctuation for char in clean_name)

                # Condition for stopping: looks like a sentence (more than 2 words and with punctuation)
                if word_count > 2 and has_punctuation:
                    # Found a sentence-like line, indicating the end of the roster list.
                    break

                # If it's not a recognized player and not a stopper, add to unrecognized list
                unrecognized_player_names.append(clean_name)
                continue # Continue to the next line, skipping this unrecognized entry for now

    if not player_names and not unrecognized_player_names:
        raise ValueError("No player names (recognized or unrecognized) extracted from the EML file.")

    return player_names, unrecognized_player_names

def assign_player(team, player, pos):
    team[pos].append(player)
    team["total"] += player.rank


# -----------------------------
# Core Optimizer
# -----------------------------

def attempt_build(players_list):

    random.shuffle(players_list)

    team_a = empty_team()
    team_b = empty_team()

    for player in players_list:

        options = []

        for team in [team_a, team_b]:

            if player.position == "F":
                if can_add(team, "F"):
                    options.append((team, "F"))

            elif player.position == "D":
                if can_add(team, "D"):
                    options.append((team, "D"))

            else:  # FLEX
                if can_add(team, "F"):
                    options.append((team, "F"))
                if can_add(team, "D"):
                    options.append((team, "D"))

        if not options:
            options = [
                (team_a, "F"),
                (team_b, "F"),
                (team_a, "D"),
                (team_b, "D")
            ]

        best_move = (None, None)
        best_diff = float("inf")

        for team, pos in options:

            proj_a = team_a["total"]
            proj_b = team_b["total"]

            if team == team_a:
                proj_a += player.rank
            else:
                proj_b += player.rank

            diff = abs(proj_a - proj_b)

            if diff < best_diff:
                best_diff = diff
                best_move = (team, pos)

        assign_player(best_move[0], player, best_move[1])

    return team_a, team_b


# -----------------------------
# Run Full Build
# -----------------------------

def build_teams(roster_eml_path, player_data_xlsx_path):
    """
    Builds balanced teams by extracting player names from an EML roster file
    and looking up their rank/position from an Excel player data file.
    """
    # 2. Read player data (ranks and positions) from the Excel file
    df_players = pd.read_excel(player_data_xlsx_path, dtype={"Name": str, "Rank": int, "Position": str})

    required_columns_xlsx = {"Name", "Rank", "Position"}

    if not required_columns_xlsx.issubset(df_players.columns):
        raise ValueError(
            f"Player data Excel file must contain '{', '.join(required_columns_xlsx)}' columns."
        )

    # Create a lookup dictionary for player data for efficient access
    player_data_lookup = {
        cast(str, row["Name"]).strip().upper(): {
            "rank": int(cast(int, row["Rank"])),
            "position": cast(str, row["Position"]).strip().upper(),
        }
        for _, row in df_players.iterrows()
    }

    # 1. Get player names from EML roster file, separating recognized from unrecognized
    eml_recognized_player_names, eml_unrecognized_player_names = extract_players_from_eml(roster_eml_path, player_data_lookup)
    if not eml_recognized_player_names and not eml_unrecognized_player_names:
        raise ValueError("No player names (recognized or unrecognized) extracted from the EML roster.")

    # --- Phase 1: Prepare recognized and unrecognized player objects ---
    recognized_players_for_tonight = []
    # Create Player objects for recognized players
    for name_from_eml in eml_recognized_player_names:
        normalized_name_eml = name_from_eml.strip().upper()
        if normalized_name_eml in player_data_lookup:
            data = player_data_lookup[normalized_name_eml]
            recognized_players_for_tonight.append(
                Player(
                    name=name_from_eml, # Use original casing for display purposes
                    rank=data["rank"],
                    position=data["position"],
                    is_unrecognized=False
                )
            )
        else:
            # This should ideally not happen if extract_players_from_eml works as expected
            # and correctly filters names against player_data_lookup.
            raise ValueError(
                f"Player '{name_from_eml}' from EML roster was not found in the player data Excel file (unexpected error, "
                "should have been caught as unrecognized)."
            )

    unrecognized_player_objects = []
    # Create Player objects for unrecognized players with default rank 0 and '?' position
    for name_from_eml in eml_unrecognized_player_names:
        unrecognized_player_objects.append(
            Player(name=name_from_eml, rank=0, position="?", is_unrecognized=True)
            )

    if len(recognized_players_for_tonight) + len(unrecognized_player_objects) < 10:
        raise ValueError("Not enough selected players for team balancing (minimum 10 needed).")

    # --- Phase 2: Balance recognized players based on skill ---

    # Prepare recognized players for the balancing algorithm
    recognized_players_to_balance = recognized_players_for_tonight.copy()
    median_recognized_player_removed = None

    # If there's an odd number of recognized players, remove a median-ranked player
    if len(recognized_players_to_balance) % 2 != 0:
        recognized_players_to_balance.sort(key=lambda p: p.rank)
        median_index = len(recognized_players_to_balance) // 2
        median_recognized_player_removed = recognized_players_to_balance.pop(median_index)

    best_recognized_teams = (None, None)
    best_diff_recognized = float("inf")

    # Run optimizer for recognized players only
    for _ in range(ITERATIONS):
        # Create a copy to ensure each iteration starts with the same player pool
        a_rec, b_rec = attempt_build(recognized_players_to_balance.copy())
        diff_rec = abs(a_rec["total"] - b_rec["total"])

        if diff_rec < best_diff_recognized:
            best_diff_recognized = diff_rec
            best_recognized_teams = (deepcopy(a_rec), deepcopy(b_rec))

        if best_diff_recognized == 0:
            break

    final_team_a = best_recognized_teams[0]
    final_team_b = best_recognized_teams[1]

    # Re-add the median recognized player if one was removed
    if median_recognized_player_removed:
        if random.choice([True, False]):
            chosen_team = final_team_a
        else:
            chosen_team = final_team_b

        # Add the player to their primary position or 'F' if FLEX
        if median_recognized_player_removed.position == "D" and len(chosen_team["D"]) < DEFENCE_TARGET:
            chosen_team["D"].append(median_recognized_player_removed)
        else:  # "F", "FLEX", or if D is full
            chosen_team["F"].append(median_recognized_player_removed)

        chosen_team["total"] += median_recognized_player_removed.rank # Update total rank

    # --- Phase 3: Distribute unrecognized players to balance total player count ---

    # Shuffle unrecognized players to ensure random distribution if counts are equal
    random.shuffle(unrecognized_player_objects)

    for unrec_player in unrecognized_player_objects:
        len_a = len(final_team_a["F"]) + len(final_team_a["D"])
        len_b = len(final_team_b["F"]) + len(final_team_b["D"])

        # Prioritize adding to the team with fewer players overall
        if len_a <= len_b:
            # Try to add to 'F' first if not full based on target, else 'D', else 'F' over capacity
            if len(final_team_a["F"]) < FORWARDS_TARGET:
                final_team_a["F"].append(unrec_player)
            elif len(final_team_a["D"]) < DEFENCE_TARGET:
                final_team_a["D"].append(unrec_player)
            else: # Both positions might be 'full' according to target, just add to 'F' as a default
                final_team_a["F"].append(unrec_player)
        else: # Team B has fewer players
            if len(final_team_b["F"]) < FORWARDS_TARGET:
                final_team_b["F"].append(unrec_player)
            elif len(final_team_b["D"]) < DEFENCE_TARGET:
                final_team_b["D"].append(unrec_player)
            else:
                final_team_b["F"].append(unrec_player)
        # IMPORTANT: Do NOT update team["total"] for unrecognized players as their rank is 0.
        # The total rank difference (best_diff_recognized) should only reflect recognized players.

    return final_team_a, final_team_b, best_diff_recognized, eml_unrecognized_player_names


# -----------------------------
# Export Workbook
# -----------------------------

def export_workbook(team_a, team_b, diff, unrecognized_players_list):

    def build_df(team, jersey):

        rows = []

        for pos in ["F", "D"]:
            for p in team[pos]: # p is a Player object
                rows.append({
                    "Name": p.name,
                    "Rank": p.rank,
                    "Position": "?" if p.is_unrecognized else p.position, # Output '?' for unrecognized players
                    "Jersey": jersey
                })

        return pd.DataFrame(rows)

    df_light = build_df(team_a, TEAM_A_JERSEY)
    df_dark = build_df(team_b, TEAM_B_JERSEY)

    summary = {
        "Metric": [
            "Light Team Total Rank",
            "Dark Team Total Rank",
            "Skill Difference"
        ],
        "Value": [
            team_a["total"],
            team_b["total"],
            diff
        ]
    }

    df_summary = pd.DataFrame(summary)

    # Add unrecognized players to the summary if any
    if unrecognized_players_list:
        missing_players_str = ", ".join(unrecognized_players_list)
        df_summary = pd.concat([
            df_summary,
            pd.DataFrame({
                "Metric": ["Unrecognized Players"],
                "Value": [missing_players_str]
            })
        ], ignore_index=True)
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:

        df_light.to_excel(writer, sheet_name=TEAM_A_NAME, index=False)
        df_dark.to_excel(writer, sheet_name=TEAM_B_NAME, index=False)
        df_summary.to_excel(writer, sheet_name="Summary", index=False)

        # ----------------------------------------------------
        # New "Teams" sheet with player names and styling
        # ----------------------------------------------------
        ws_teams = writer.book.create_sheet("Teams")

        # Set headers
        ws_teams['A1'] = TEAM_A_NAME
        ws_teams['B1'] = TEAM_B_NAME

        # Define styles
        header_font = Font(name='Comic Sans MS', size=28, bold=True)
        player_font = Font(name='Comic Sans MS', size=22)
        center_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Apply header styles
        ws_teams['A1'].font = header_font
        ws_teams['A1'].alignment = center_alignment
        ws_teams['B1'].font = header_font
        ws_teams['B1'].alignment = center_alignment

        # Collect and sort player names for each team
        all_players_a = sorted([p.name for p in team_a["F"] + team_a["D"]])
        all_players_b = sorted([p.name for p in team_b["F"] + team_b["D"]])

        # Populate player names and apply styles
        max_rows = max(len(all_players_a), len(all_players_b)) + 1 # +1 for header row

        for i, name in enumerate(all_players_a):
            cell_a = ws_teams.cell(row=i+2, column=1, value=name)
            cell_a.font = player_font

        for i, name in enumerate(all_players_b):
            cell_b = ws_teams.cell(row=i+2, column=2, value=name)
            cell_b.font = player_font

        # Apply borders to the entire populated range
        for row in ws_teams.iter_rows(min_row=1, max_row=max_rows, min_col=1, max_col=2):
            for cell in row:
                cell.border = thin_border

        # Adjust column widths for better readability (auto-fit)
        column_widths = {}
        for row in ws_teams.iter_rows(min_row=1, max_row=max_rows, min_col=1, max_col=2):
            for cell in row:
                if cell.value:
                    # Estimate width based on character count and font size.
                    # These factors are heuristics to approximate visible width in Excel.
                    col_letter = cell.column_letter
                    if cell.row == 1: # Header row (28pt bold)
                        width_estimate = len(str(cell.value)) * 3.0
                    else: # Player names (22pt)
                        width_estimate = len(str(cell.value)) * 2.5

                    current_max = column_widths.get(col_letter, 0)
                    column_widths[col_letter] = max(current_max, width_estimate)

        for col_letter, width in column_widths.items():
            # Add a small padding to ensure text isn't cramped
            ws_teams.column_dimensions[col_letter].width = width + 2


# -----------------------------
# GUI
# -----------------------------

def update_generate_button_state():
    """Enables the generate button only when both EML and XLSX files are selected."""
    eml_selected = eml_file_label.cget("text") != "No EML file selected"
    xlsx_selected = player_data_file_label.cget("text") != "No player data file selected"
    if eml_selected and xlsx_selected:
        generate_button.config(state="normal")
    else:
        generate_button.config(state="disabled")

def select_eml_file():
    """Opens a file dialog for selecting the EML roster file."""
    file_path = filedialog.askopenfilename(
        title="Select Roster EML File",
        filetypes=[("EML Files", "*.eml")]
    )
    if file_path:
        eml_file_label.config(text=file_path)
        update_generate_button_state()

def select_player_data_file():
    """Opens a file dialog for selecting the player data Excel file."""
    file_path = filedialog.askopenfilename(
        title="Select Player Data Excel File",
        filetypes=[("Excel Files", "*.xlsx")]
    )
    if file_path:
        player_data_file_label.config(text=file_path)
        update_generate_button_state()


def generate_teams():
    """
    Triggers the team generation process using the selected EML and Excel files.
    Handles errors and displays success/failure messages.
    """
    try:
        eml_file_path = eml_file_label.cget("text")
        player_data_xlsx_path = player_data_file_label.cget("text")

        if eml_file_path == "No EML file selected" or player_data_xlsx_path == "No player data file selected":
            raise ValueError("Please select both the EML roster file and the player data Excel file.")

        team_a, team_b, diff, unrecognized_players = build_teams(eml_file_path, player_data_xlsx_path)
        export_workbook(team_a, team_b, diff, unrecognized_players)

        status_message = f"✔ Teams Created Successfully!\nSkill Difference: {diff}"
        if unrecognized_players:
            status_message += f"\n(Some players were unrecognized - see Summary sheet)"
            status_label.config(text=status_message, fg="orange") # Use orange to indicate warnings
            messagebox.showwarning(
                "Unrecognized Players",
                "The following players were found in the roster but not recognized "
                "from your player data file and were added with a rank of 0 and position '?':\n\n" +
                "\n".join(unrecognized_players) +
                "\n\nPlease ensure their names are correctly listed in your player data Excel file if you wish them to be fully recognized."
            )
        else:
            status_label.config(text=status_message, fg="green")

        messagebox.showinfo(
            "Success", f"Workbook created:\n{OUTPUT_FILE}"
        )

    except Exception as e:
        status_label.config(text="❌ Error Occurred", fg="red")
        messagebox.showerror("Error", str(e))


# -----------------------------
# Build Window
# -----------------------------

root = tk.Tk()
root.title("Hockey Team Balancer")
# Increase window size to accommodate new elements
root.geometry("460x400") # Original was 460x300
root.resizable(False, False)

title = tk.Label(root, text="🏒 Hockey Team Balancer", font=("Arial", 16, "bold"))
title.pack(pady=10)

# EML File Selection Components
select_eml_button = tk.Button(root, text="Select Roster EML File", command=select_eml_file, width=30)
select_eml_button.pack(pady=5)
eml_file_label = tk.Label(root, text="No EML file selected", wraplength=420)
eml_file_label.pack(pady=2)

# Player Data XLSX File Selection Components
select_player_data_button = tk.Button(root, text="Select Player Data Excel File", command=select_player_data_file, width=30)
select_player_data_button.pack(pady=5)
player_data_file_label = tk.Label(root, text="No player data file selected", wraplength=420)
player_data_file_label.pack(pady=2)

generate_button = tk.Button(
    root,
    text="Generate Balanced Teams",
    command=generate_teams,
    width=25,
    state="disabled"
)
generate_button.pack(pady=15)

status_label = tk.Label(root, text="", font=("Arial", 11))
status_label.pack(pady=10)

root.mainloop()
